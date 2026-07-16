"""Shared engine for the stability harnesses: judging, clustering, reporting.

Both stability scripts (``publishing_stability.py``, ``critique_stability.py``)
measure run-to-run reproducibility the same way: candidate finding pairs are scored
as same-problem (lexical Jaccard gate; an LLM judge only for the ambiguous middle
band), the passing pairs become edges, and each connected component is one distinct
issue whose support is the number of runs it appears in. What differs per checker —
how findings are loaded and which pairs are candidates — stays in the scripts;
everything checker-agnostic lives here.
"""

import asyncio
import re
import sys
import time
from collections.abc import Awaitable, Callable
from dataclasses import dataclass
from datetime import UTC, datetime
from itertools import combinations
from pathlib import Path
from typing import Any, Protocol

import pydantic
import pydantic_ai
from pydantic_ai.settings import ModelSettings

from scripts.console import dim, green, red, yellow
from scripts.evaluations_common import issue_jaccard
from scripts.evaluations_runs import default_output_dir

# A judge: given two issue texts, score 0.0-1.0 how much they are the same problem.
JudgeFn = Callable[[str, str], Awaitable[float]]

# Tiering thresholds for the same-problem decision. Above HIGH the lexical overlap
# is decisive (match without judging); at or below LOW it is decisive the other way;
# the band between is sent to the judge. THRESHOLD is the same-problem cut-off applied
# to the resulting score when drawing a clustering edge.
DEFAULT_HIGH_JACCARD = 0.6
DEFAULT_LOW_JACCARD = 0.1
DEFAULT_THRESHOLD = 0.5
# Max judge (LLM) calls in flight at once. 1 = sequential. The unit bounded is the
# pairwise comparison; raising it parallelises the judging within a candidate batch.
DEFAULT_CONCURRENCY = 1
# Runs to generate from --document when --runs is not given.
DEFAULT_RUNS = 5

# Judge-call resilience. Bedrock throttling (ThrottlingException) is retried by
# boto3 internally with backoff, so it surfaces as slow calls rather than errors;
# we both retry exhausted throttles ourselves and warn on either signal so a long
# run's slowness is attributable to rate limiting rather than guessed at.
SLOW_CALL_WARN_S = 12.0
MAX_JUDGE_ATTEMPTS = 6
# Base of the exponential backoff between throttle retries: 5, 10, 20, 40, 80s.
THROTTLE_BACKOFF_S = 5.0
_THROTTLE_MARKERS = (
    "throttl",
    "too many requests",
    "toomanyrequests",
    "rate exceeded",
    "slowdown",
    "429",
)

JUDGE_SYSTEM_PROMPT = (
    "You compare two issue descriptions, each written by an automated document "
    "checker on a separate run over the same document. They are peers — neither is "
    "reference, and their A/B order is arbitrary. Respond with a JSON object "
    "{reason, score}: score 1.0 if they identify the same underlying problem, 0.0 "
    "if unrelated, intermediate for partial overlap. Judge the substance of the "
    "problem, not the wording."
)


class SameProblemGrade(pydantic.BaseModel):
    """The judge's verdict on whether two issue descriptions are the same problem."""

    reason: str
    score: float


# Own judge agent (the library's judges discard usage; this keeps the run result so
# tokens can be counted). The Bedrock model is supplied per run, as the checker does.
_judge_agent = pydantic_ai.Agent(
    output_type=SameProblemGrade,
    system_prompt=JUDGE_SYSTEM_PROMPT,
)


@dataclass
class JudgeUsage:
    """Running token totals across the judge calls made during a comparison."""

    calls: int = 0
    input_tokens: int = 0
    output_tokens: int = 0


def support_colour(support: int, n_runs: int, text: str) -> str:
    """Green when in every run, red when in only one, yellow in between."""
    if support <= 1:
        return red(text)
    return green(text) if support >= n_runs else yellow(text)


def warn(message: str) -> None:
    """Emit a diagnostic to stderr, kept off the stdout report."""
    print(f"[stability] {message}", file=sys.stderr, flush=True)


def progress(message: str) -> None:
    """Emit a dim progress line to stderr, kept off the stdout report."""
    print(dim(f"[stability] {message}"), file=sys.stderr, flush=True)


def _is_throttle(error: BaseException) -> bool:
    """Whether an exception looks like a rate-limit / throttling response."""
    text = f"{type(error).__name__} {error}".lower()
    return any(marker in text for marker in _THROTTLE_MARKERS)


def truncate(text: str, width: int = 100) -> str:
    """Collapse whitespace and cut to ``width`` with an ellipsis."""
    collapsed = " ".join(text.split())
    return collapsed if len(collapsed) <= width else f"{collapsed[: width - 1]}…"


def run_label(path: Path) -> str:
    """A compact run name: the trailing ``runNN`` token if present, else the stem."""
    stem = path.stem
    _, marker, tail = stem.rpartition("run")
    if marker and tail.isdigit():
        return f"run{tail}"
    return stem


def dotted_tuple(number: str) -> tuple[int, ...]:
    """A dotted-decimal section number as an int tuple, so '10' sorts after '2'."""
    return tuple(int(part) for part in number.split("."))


async def same_problem_score(
    first: str, second: str, judge: JudgeFn, low: float, high: float
) -> float:
    """Tiered 0.0-1.0 score that two issue texts are the same problem.

    Lexical Jaccard decides the clear cases (keeping the judge — itself a variance
    source — out of them); the judge is consulted only for the ambiguous band. The
    pair is canonicalised before judging so the score is symmetric by construction.
    """
    jaccard = issue_jaccard(first, second)
    if jaccard >= high:
        return 1.0
    if jaccard <= low:
        return 0.0
    canonical_first, canonical_second = sorted((first, second))
    return await judge(canonical_first, canonical_second)


def connected_components(size: int, edges: list[tuple[int, int]]) -> list[list[int]]:
    """Group ``range(size)`` into connected components given undirected ``edges``."""
    parent = list(range(size))

    def find(node: int) -> int:
        while parent[node] != node:
            parent[node] = parent[parent[node]]
            node = parent[node]
        return node

    for left, right in edges:
        parent[find(left)] = find(right)

    groups: dict[int, list[int]] = {}
    for node in range(size):
        groups.setdefault(find(node), []).append(node)
    return list(groups.values())


def bounded_judge(judge: JudgeFn, concurrency: int) -> JudgeFn:
    """A judge that admits at most ``concurrency`` calls at once via a semaphore.

    Only the (slow) judge call is bounded; the Jaccard gate runs unbounded, so a slot
    is held for an actual LLM call, never for a pair the lexical gate settles.
    """
    semaphore = asyncio.Semaphore(concurrency)

    async def bounded(first: str, second: str) -> float:
        async with semaphore:
            return await judge(first, second)

    return bounded


async def score_edges(
    texts: list[str],
    pairs: list[tuple[int, int]],
    judge: JudgeFn,
    low: float,
    high: float,
    threshold: float,
) -> list[tuple[int, int]]:
    """Same-problem edges among the candidate ``pairs`` (indices into ``texts``).

    The pairs are independent, so they are scored concurrently; how many judge
    calls actually overlap is bounded by the semaphore inside ``judge``.
    """
    scores = await asyncio.gather(
        *(same_problem_score(texts[i], texts[j], judge, low, high) for i, j in pairs)
    )
    return [
        pair for pair, score in zip(pairs, scores, strict=True) if score >= threshold
    ]


class ClusterLike(Protocol):
    """What the agreement/histogram aggregates need to know about a cluster."""

    @property
    def runs(self) -> set[str]: ...

    @property
    def support(self) -> int: ...


def pairwise_agreements(
    runs: list[str], clusters: list[ClusterLike]
) -> dict[tuple[str, str], float]:
    """Soft-Dice agreement for every unordered run pair, from the cluster supports.

    Each run is the set of issue-clusters it touched; agreement is
    ``2 * shared / (|i| + |j|)``. Symmetric, and 1.0 for two runs that both raised
    nothing.
    """
    touched = {run: sum(run in cluster.runs for cluster in clusters) for run in runs}
    agreements: dict[tuple[str, str], float] = {}
    for first, second in combinations(runs, 2):
        shared = sum(
            1
            for cluster in clusters
            if first in cluster.runs and second in cluster.runs
        )
        denominator = touched[first] + touched[second]
        agreements[first, second] = 2 * shared / denominator if denominator else 1.0
    return agreements


def support_histogram(clusters: list[ClusterLike], n_runs: int) -> dict[int, int]:
    """How many clusters have each support level 1..N."""
    histogram = dict.fromkeys(range(1, n_runs + 1), 0)
    for cluster in clusters:
        histogram[cluster.support] = histogram.get(cluster.support, 0) + 1
    return histogram


def match_report_path(
    document: str | None,
    out_dir: str | None,
    paths: list[Path],
    run_file_suffix: re.Pattern[str],
    checker: str,
) -> Path:
    """Where to write the match report: '<input>-<checker>-match-report-<ts>.xlsx'.

    The checker infix keeps one checker's reports distinguishable from another's
    for the same document, matching the run captures' naming. Written to
    ``out_dir``, or ``default_output_dir()`` when that is unset. The input stem
    is the document name for --document, otherwise recovered from the first run
    file by stripping its batch/run suffix (``run_file_suffix``, checker-specific).
    """
    if document is not None:
        stem = Path(document).stem
    else:
        stem = run_file_suffix.sub("", paths[0].stem)
    directory = Path(out_dir) if out_dir else default_output_dir()
    timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    return directory / f"{stem}-{checker}-match-report-{timestamp}.xlsx"


def write_match_sheet(
    sheet: Any, runs: list[str], rows: list[tuple[str, float, list[str]]]
) -> None:
    """Fill one match-report worksheet: a row per cluster, a column per run.

    ``rows`` is (section text, match fraction, one cell of issue text per run).
    Formatting applied:
    - Font: Aptos Narrow 11pt throughout.
    - Row 1 headers: bold with a bottom border only.
    - Section column (A): text number format so "4.1" is never coerced to a decimal.
    - Match fraction column (B): float stored as 0% percentage with Excel's standard
      red→white→blue 3-colour scale (F8696B / FCFCFF / 5A8AC6).
    - Run columns (C+): 60 character units wide, word-wrapped, top-aligned.
    """
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    base_font = Font(name="Aptos Narrow", size=11)
    header_font = Font(name="Aptos Narrow", size=11, bold=True)
    header_border = Border(bottom=Side(style="thin"))
    top = Alignment(vertical="top")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col, text in enumerate(["section", "match", *runs], 1):
        cell = sheet.cell(row=1, column=col, value=text)
        cell.font = header_font
        cell.border = header_border

    for row_idx, (section, fraction, cells) in enumerate(rows, start=2):
        sec = sheet.cell(row=row_idx, column=1, value=section)
        sec.number_format = "@"
        sec.alignment = top
        sec.font = base_font

        frac = sheet.cell(row=row_idx, column=2, value=fraction)
        frac.number_format = "0%"
        frac.alignment = top
        frac.font = base_font

        for col_idx, value in enumerate(cells, start=3):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
            cell.font = base_font

    if rows:
        # Excel's built-in Red–White–Blue preset: low=flaky (red), high=stable (blue).
        sheet.conditional_formatting.add(
            f"B2:B{1 + len(rows)}",
            ColorScaleRule(
                start_type="min",
                start_color="FFF8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFFCFCFF",
                end_type="max",
                end_color="FF5A8AC6",
            ),
        )

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 7
    for i in range(3, 3 + len(runs)):
        sheet.column_dimensions[get_column_letter(i)].width = 60


def make_bedrock_judge(model: Any) -> tuple[JudgeFn, JudgeUsage]:
    """A Bedrock-backed judge (temperature 0) and the usage it accumulates.

    Each call's input/output tokens are summed into the returned ``JudgeUsage`` so
    the cost of the similarity checks can be reported. Only middle-band pairs reach
    here; Jaccard-resolved pairs cost nothing.
    """
    usage = JudgeUsage()

    async def run_once(prompt: str) -> tuple[Any, float]:
        """Run the judge once, retrying explicit throttles with backoff."""
        for attempt in range(1, MAX_JUDGE_ATTEMPTS + 1):
            start = time.perf_counter()
            try:
                result = await _judge_agent.run(
                    prompt, model=model, model_settings=ModelSettings(temperature=0.0)
                )
            except Exception as error:  # noqa: BLE001 — classify, warn, then re-raise
                if not _is_throttle(error):
                    raise
                if attempt == MAX_JUDGE_ATTEMPTS:
                    warn(
                        f"still rate limited after {MAX_JUDGE_ATTEMPTS} attempts on "
                        f"judge call {usage.calls + 1}; giving up"
                    )
                    raise
                delay = THROTTLE_BACKOFF_S * 2 ** (attempt - 1)
                warn(
                    f"rate limited on judge call {usage.calls + 1} "
                    f"(attempt {attempt}/{MAX_JUDGE_ATTEMPTS}); backing off {delay:.0f}s"
                )
                await asyncio.sleep(delay)
            else:
                return result, time.perf_counter() - start
        msg = "unreachable: judge retry loop exhausted without return or raise"
        raise RuntimeError(msg)

    async def judge(first: str, second: str) -> float:
        prompt = f"<IssueA>\n{first}\n</IssueA>\n<IssueB>\n{second}\n</IssueB>"
        result, elapsed = await run_once(prompt)
        run_usage = result.usage
        usage.calls += 1
        usage.input_tokens += run_usage.input_tokens or 0
        usage.output_tokens += run_usage.output_tokens or 0
        # A slow call with no exception means boto3 is silently retrying a throttle.
        if elapsed > SLOW_CALL_WARN_S:
            warn(
                f"judge call {usage.calls} took {elapsed:.0f}s — likely being "
                "throttled (boto3 is backing off internally)"
            )
        return result.output.score

    return judge, usage
